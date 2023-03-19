import openpyxl as xl
import os

# Структура передаваемых данных:
# [{}, {}, {}, {}] -  для каждого сайта
# {"наименование позиции": "", "категория": "батарея/микроконтроллер/т.д.", прочие характеристики} 
class XLWriter:
	def letterExtend(self, index):
		"""Перевод числового индекса столбца в Excel-евский буквенный формат.
		Индексация начинается с 1"""
		alphabet = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
		out = ""
		while index > 0:
			out += alphabet[(index-1)%26]
			index = (index-1)//26

		return out[::-1]

	def findLastCol(self):
		i = 0
		for row in self.sheet.iter_cols(min_row=0, max_row=1, min_col=0, max_col=26*26): # как-то по-другому задать макс.индекс для столбца
				for cell in row:
					if cell.value is None:
						self.colindex = i
						print(i)
						return
					else:
						self.col_list[cell.value] = self.letterExtend(i+1)
						print(i)
					i += 1
		return -1

	def __init__(self, data):
		self.data = data
		self.col_list = {}	# Словарь адресов столбцов по их названиям
		self.colindex = 0
		# self.greenc = xl.styles.PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
		self.greenc = xl.styles.PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
		# self.whitec = xl.styles.PatternFill(start_color="000000", end_color="000000", fill_type="solid")

		if not os.path.isfile("output.xlsx"): # подтягивать название файла из конфига?
			self.wb = xl.Workbook()
			self.sheet = self.wb.active
		else:
			self.wb = xl.load_workbook("output.xlsx")
			self.sheet = self.wb.active
			self.findLastCol()
			for row in self.sheet.iter_cols(min_row=0, max_row=self.sheet.max_row, min_col=0, max_col=self.sheet.max_column):
				for cell in row:
					cell.fill = xl.styles.PatternFill(fill_type=None)

		for dct in data:
			for title in dct:
				if title not in self.col_list:
					self.col_list[title] = self.letterExtend(self.colindex+1)
					self.sheet[self.letterExtend(self.colindex+1)+"1"] = title
					self.sheet[self.letterExtend(self.colindex+1)+"1"].fill = self.greenc
					self.colindex += 1

		print(self.col_list)

	def populateSheet(self):
		"""Заполняет таблицу данными"""
		for dct in data:
			row = {}
			k = self.sheet.max_row
			for title in dct:
				row[self.col_list[title]] = dct[title]
				# self.sheet[self.col_list[title]+str(k)+"1"].fill = self.greenc

			self.sheet.append(row)
			for col in self.sheet.iter_cols(min_col=0, max_col=self.sheet.max_column, 
							min_row=self.sheet.max_row, max_row=self.sheet.max_row):
					for cell in col:
						cell.fill = self.greenc

	def save(self):
		"""Сохраняет таблицу"""
		self.wb.save(filename="output.xlsx")


if __name__ == "__main__":
	# Пример входных данных
	data = [{"Наименование": "Аккум #1", "Категория": "Батарея", "Разряд тока": "5 A"},
			{"Наименование": "КТ-МТ 9", "Категория": "Контроллер мотора", "Рабочий ток": "250 мА", "Пиковый ток": "500 мА"},
			{"Наименование": "Ктрл 56", "Категория": "Микроконтроллер", "Рабочий ток": "200 мА"},]
	# data = [{"Наименование": "Вундервафля", "Категория": "Лидар", "Частота импульсов": "100 Гц"}]

	w = XLWriter(data)
	w.populateSheet()
	w.save()